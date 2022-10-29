import logging
import time
from datetime import datetime
import os
try:
    import docx
    import pymorphy2
    from openpyxl import load_workbook
    from number_converter import convert_price # модуль конвертации цифр в слова
except ModuleNotFoundError as ex:
    print(ex)
    time.sleep(5)
    exit()


def get_json_from_sheet() -> dict:
    """Создает json на основе данных всех листов Excel."""
    book = load_workbook('Оплата.xlsx')
    data = {}
    for sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        row_generator = sheet.values
        try:
            headers = next(row_generator)
        except StopIteration:
            logging.error(f'Лист {sheet_name} книги пустой')
            raise Exception
        data[sheet_name] = []
        for row in row_generator:
            dict_row = dict(zip(headers, row))
            data[sheet_name].append(dict_row)
    book.close()

    return data


def str_in_padej(str_nomn: str, padej: str) -> str:
    """Принимает строку str_nomn в именительном падеже.
    Переводит строку в падежь padej.
    Наименование падежей указан в библиотеке pymorphy2"""
    str_datv = ''
    try:
        morph = pymorphy2.MorphAnalyzer('data\\') # путь к папке data указывается на случай конвертации в EXE
    except:
        logging.error(f'Отсутствует папка data')
        try:
            morph = pymorphy2.MorphAnalyzer()
        except Exception as ex:
            logging.error(ex)
            raise Exception()
    for elm in str_nomn.split(' '):
        if 'ого' == elm[-3:]:
            elm_datv = elm
        else:
            elm_datv = morph.parse(elm)[0].inflect({padej}).word
        if elm[0].isupper():
            elm_datv = elm_datv.capitalize()
        str_datv += ' ' + elm_datv
    return str_datv.strip()


def get_orders(name_company: str, list_persons: list[dict]):
    """Создает приказы в формате docx.
    Помещает созданные файлы в соответствующую папку"""
    for person in list_persons:
        fio_nomn = person['ФИО']
        prize = person['Премия']
        doc_number = person['Номер приказа']
        doc_date = list(person.keys())[-1]
        fio_datv = person['ФИО в дательном падеже']
        post_datv = person['Должность в дательном падеже']

        prize_str = convert_price(prize)
        month_list = ['январь', 'февраль', 'март', 'апрель', 'май', 'июнь',
                      'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь']
        month_nomn = month_list[doc_date.month - 1]
        month_loct = str_in_padej(month_nomn, 'loct')
        params = {
            'doc_date': doc_date.strftime('%d.%m.%Y'),
            'month_nomn': month_nomn,
            'doc_number': str(doc_number),
            'month_loct': month_loct,
            'year': str(doc_date.year),
            'fio_nomn': fio_nomn,
            'fio_datv': fio_datv,
            'posts_datv': post_datv,
            'salary': f'{prize} ({prize_str})'
        }
        name_pattern_docx = f'Шаблон {name_company} (приказы).docx'

        path_dir = f'{params["month_nomn"]} {params["year"]}'
        if not os.path.exists(path_dir):
            os.mkdir(path_dir)

        path_dir = os.path.join(path_dir, name_company)
        if not os.path.exists(path_dir):
            os.mkdir(path_dir)

        path_dir = os.path.join(path_dir, 'Приказы')
        if not os.path.exists(path_dir):
            os.mkdir(path_dir)

        try:
            save_file(name_pattern_docx, path_dir, params)
        except PermissionError:
            continue


def get_month_nomn(month: int) -> str:
    """Возвращает название месяца на Русском по его номеру"""
    month_list = ['январь', 'февраль', 'март', 'апрель', 'май', 'июнь',
                  'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь']
    return month_list[month - 1]


def get_date_str(date: datetime) -> str:
    """Записывает дату в формате:
    '01' января 2022 г."""
    month_nomn = get_month_nomn(date.month)
    month_loct = str_in_padej(month_nomn, 'gent')
    return f'"{date.day}" {month_loct} {date.year} г.'


def get_fio_abbreviated(fio: str) -> str:
    """Из полного ФИО получает Фамилию и инициалы"""
    try:
        fio_elments = fio.split(' ')
        return f'{fio_elments[0]} {fio_elments[1][0]}. {fio_elments[2][0]}.'
    except:
        logging.error(f'ФИО для {fio} записано в неправильном формате')
        raise Exception()


def get_act(name_company, list_persons):
    """Создает акты в формате docx.
        Помещает созданные файлы в соответствующую папку"""
    for person in list_persons:
        fio_nomn = person['ФИО']
        gender = person['Пол']
        contract_number = person['Номер договора подряда']
        contract_date = person['Дата договора подряда']
        start_date = person['Дата начала работ']
        end_date = person['Дата окончания работ']
        total_price = person['Общая стоимость работ']
        final_price = person['Итого к выдаче']
        doc_number = person['Номер акта']
        doc_date = list(person.keys())[-1]

        total_price_str = convert_price(total_price)
        final_price_str = convert_price(final_price)
        if gender == 'муж':
            gender = 'ый'
        elif gender == 'жен':
            gender = 'ая'
        else:
            logging.error(f'Пол для {fio_nomn} задан не корректно. Пропускаю создание акта')
            continue
        params = {
            'doc_date': get_date_str(doc_date),
            'fio_nomn': fio_nomn,
            'gender': gender,
            'doc_number': str(doc_number),
            'contract_number': str(contract_number),
            'contract_date': get_date_str(contract_date),
            'end_date': get_date_str(end_date),
            'start_date': get_date_str(start_date),
            'total_price': f'{total_price} ({total_price_str})',
            'final_price': f'{final_price} ({final_price_str})',
            'month_nomn': get_month_nomn(doc_date.month),
            'year': str(doc_date.year),
            'fio_reduction': get_fio_abbreviated(fio_nomn)
        }
        name_pattern_docx = f'Шаблон {name_company} (акты).docx'

        path_dir = f'{params["month_nomn"]} {params["year"]}'
        if not os.path.exists(path_dir):
            os.mkdir(path_dir)

        path_dir = os.path.join(path_dir, name_company)
        if not os.path.exists(path_dir):
            os.mkdir(path_dir)

        path_dir = os.path.join(path_dir, 'Акты')
        if not os.path.exists(path_dir):
            os.mkdir(path_dir)
        try:
            save_file(name_pattern_docx, path_dir, params)
        except PermissionError:
            continue


def save_file(name_pattern_docx: str, path_dir: str, params: dict):
    """Формирует docx на основе шаблона name_pattern_docx,
    с учётом параметров params, и сохраняет по пути path_dir"""
    try:
        doc = docx.Document(name_pattern_docx)
    except docx.opc.exceptions.PackageNotFoundError:
        logging.error(f'{name_pattern_docx} не найден')
        raise Exception()
    for paragraph in doc.paragraphs:
        paragraph.text = paragraph.text.replace('{', '').replace('}', '')
        for key, value in params.items():
            if f'{key}' in paragraph.text:
                paragraph.text = paragraph.text.replace(f'{key}', value)
    name_file = f'{params["doc_number"]} {params["fio_nomn"]}.docx'
    path_file = os.path.join(path_dir, name_file)
    try:
        doc.save(path_file)
    except PermissionError:
        logging.error(f'Файл {path_file} заблокирован пользователем. Закройте файл')
        raise PermissionError()
    logging.info(f'Создан файл: {path_file}')


def main():
    data = get_json_from_sheet()
    for key, value in data.items():
        name_company = ' '.join(key.split(' ')[:-1])
        if 'приказы' in key:
            try:
                get_orders(name_company, value)
            except:
                logging.error(f'Не удалось создать приказы для {name_company}')
                continue
        elif 'акты' in key:
            try:
                get_act(name_company, value)
            except:
                logging.error(f'Не удалось создать акты для {name_company}')
                continue


def log():
    """Логирование скрипта в консоль и файл"""
    logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
    root_logger = logging.getLogger()
    root_logger.setLevel(logging.INFO)
    handler = logging.FileHandler('Протокол.txt', 'a', 'utf-8')
    handler.setFormatter(logging.Formatter('%(asctime)s %(levelname)s %(message)s'))
    handler.setLevel(logging.INFO)
    root_logger.addHandler(handler)
    logging.getLogger("pymorphy2").setLevel(logging.ERROR)


if __name__ == '__main__':
    log()
    logging.info('Приступаю к созданию документов')
    try:
        main()
        logging.info('Создание документов окончено')
    except:
        logging.error('ОШИБКА!!!')
    time.sleep(2)
